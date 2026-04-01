"""Читання та запис Excel для агента ТТН."""
import csv
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Канонічні назви полів після нормалізації
COL_ID = "id"
COL_PHONE = "phone"
COL_CITY = "city"
COL_WAREHOUSE = "warehouse"
COL_NAME = "name"
COL_PRODUCT = "product"
COL_QTY = "qty"

# Варіанти заголовків у Бітрікс24 (lowercase, strip)
_HEADER_MAP = {
    "id": COL_ID,
    "нова пошта - номер телефону отримувача": COL_PHONE,
    "нова пошта - місто отримувача": COL_CITY,
    "нова пошта - номер відділення (число)": COL_WAREHOUSE,
    "нова пошта - піб отримувача": COL_NAME,
    "товар": COL_PRODUCT,
    "кількість": COL_QTY,
}


def _normalize_header(h: str) -> str:
    return h.strip().lower() if h else ""


def read_bitrix_export(path: str | Path) -> list[dict]:
    """Читає Excel-експорт з Бітрікс24 (.xlsx або .xls / HTML-таблиця).
    Повертає список нормалізованих рядків."""
    path = Path(path)

    # Спочатку пробуємо openpyxl (xlsx)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        # Fallback: pandas (підтримує .xls через xlrd і HTML-таблиці через lxml)
        import pandas as pd
        try:
            df = pd.read_excel(path, dtype=str, header=0)
        except Exception:
            df = pd.read_html(str(path), header=0)[0].astype(str)
        df = df.fillna("")
        rows = [tuple(df.columns)] + list(df.itertuples(index=False, name=None))

    if not rows:
        return []

    raw_headers = [str(h) if h is not None else "" for h in rows[0]]
    col_map: dict[int, str] = {}
    for i, rh in enumerate(raw_headers):
        norm = _normalize_header(rh)
        if norm in _HEADER_MAP:
            col_map[i] = _HEADER_MAP[norm]

    result = []
    for row in rows[1:]:
        if all(v is None or str(v).strip() in ("", "None", "nan") for v in row):
            continue
        rec: dict = {}
        for i, field in col_map.items():
            val = row[i] if i < len(row) else None
            rec[field] = str(val).strip() if val is not None else ""
            if rec[field] in ("None", "nan"):
                rec[field] = ""
        if rec.get(COL_ID):
            result.append(rec)

    return result


def write_missing(rows: list[dict], output_dir: Path) -> Path:
    """Записує рядки без НП-даних у missing_YYYYMMDD.xlsx."""
    today = date.today().strftime("%Y%m%d")
    path = output_dir / f"missing_{today}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Без НП-даних"

    headers = [COL_ID, COL_NAME, COL_PRODUCT, COL_QTY, COL_PHONE, COL_CITY, COL_WAREHOUSE]
    _write_header_row(ws, 1, headers)
    for r, row in enumerate(rows, start=2):
        for c, h in enumerate(headers, start=1):
            ws.cell(r, c, row.get(h, ""))

    wb.save(path)
    return path


def write_ttn_results(rows: list[dict], output_dir: Path) -> Path:
    """Записує ttn_results_YYYYMMDD.xlsx."""
    today = date.today().strftime("%Y%m%d")
    path = output_dir / f"ttn_results_{today}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ТТН"

    headers = ["ТТН", "ID_угод", "Телефон", "ПІБ", "Місто", "Відділення", "Товари", "Статус"]
    _write_header_row(ws, 1, headers)
    for r, row in enumerate(rows, start=2):
        ws.cell(r, 1, row.get("ttn", ""))
        ws.cell(r, 2, row.get("ids", ""))
        ws.cell(r, 3, row.get("phone", ""))
        ws.cell(r, 4, row.get("name", ""))
        ws.cell(r, 5, row.get("city", ""))
        ws.cell(r, 6, row.get("warehouse", ""))
        ws.cell(r, 7, row.get("products", ""))
        ws.cell(r, 8, row.get("status", ""))

    wb.save(path)
    return path


def write_ttn_per_deal(
    all_rows: list[dict],
    id_to_ttn: dict[str, str],
    output_dir: Path,
) -> Path:
    """Плаский файл: кожен рядок = одна угода + її ТТН.
    Зручний для імпорту у фулфілмент.
    """
    today = date.today().strftime("%Y%m%d")
    path = output_dir / f"ttn_per_deal_{today}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ТТН по угодах"

    headers = ["ТТН", "ID угоди", "ПІБ отримувача", "Місто", "Відділення", "Телефон", "Товар", "Кількість"]
    _write_header_row(ws, 1, headers)

    # Зелений для рядків з ТТН, сірий без
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    gray_fill  = PatternFill("solid", fgColor="F2F2F2")

    r = 2
    for row in all_rows:
        deal_id = row.get(COL_ID, "")
        ttn = id_to_ttn.get(deal_id, "")
        if not ttn:
            continue  # пропускаємо без ТТН (missing або помилка)
        fill = green_fill if ttn and ttn != "DRY-RUN" else gray_fill
        values = [
            ttn,
            deal_id,
            row.get(COL_NAME, ""),
            row.get(COL_CITY, ""),
            row.get(COL_WAREHOUSE, ""),
            row.get(COL_PHONE, ""),
            row.get(COL_PRODUCT, ""),
            row.get(COL_QTY, ""),
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(r, c, v)
            cell.fill = fill
        r += 1

    # Авторозмір колонок
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(path)
    return path


def read_ttn_results(path: str | Path) -> list[dict]:
    """Читає ttn_results.xlsx."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        rec = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}
        if rec.get("ТТН"):
            result.append(rec)
    return result


def read_ttn_per_deal(path: str | Path) -> list[dict]:
    """Читає ttn_per_deal_*.xlsx для скрипту 2 (фулфілмент)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        rec = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}
        if rec.get("ТТН"):
            result.append(rec)
    return result


def write_fulfillment_orders(rows: list[dict], output_dir: Path) -> Path:
    """Записує таблицю замовлень фулфілменту НП.

    Формат (відповідає веб-формі НП):
      ТТН | Номер замовлення | Артикул | Кількість | ПІБ отримувача | Місто

    Рядки одного ТТН підсвічені одним кольором (блакитний / білий по черзі).
    """
    today = date.today().strftime("%Y%m%d")
    path = output_dir / f"fulfillment_orders_{today}.xlsx"

    headers   = ["ТТН", "Номер замовлення", "Артикул", "Кількість", "ПІБ отримувача", "Місто"]
    field_keys = ["ttn", "order_number",    "article", "qty",       "name",            "city"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Фулфілмент замовлення"
    _write_header_row(ws, 1, headers)

    # Чергуємо кольори рядків по групах ТТН
    palette = ["DCE6F1", "FFFFFF"]   # блакитний / білий
    ttn_color: dict[str, str] = {}
    color_idx = 0

    for r, row in enumerate(rows, start=2):
        ttn = row.get("ttn", "")
        if ttn not in ttn_color:
            ttn_color[ttn] = palette[color_idx % 2]
            color_idx += 1
        fill = PatternFill("solid", fgColor=ttn_color[ttn])
        for c, key in enumerate(field_keys, start=1):
            cell = ws.cell(r, c, row.get(key, ""))
            cell.fill = fill

    # Авторозмір колонок
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(path)
    return path


def _write_header_row(ws, row_num: int, headers: list[str]):
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row_num, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
